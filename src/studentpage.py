#!/usr/bin/python
# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'studentpage.ui'
#
# Created: Tue Jul 28 15:14:11 2015
#      by: PyQt4 UI code generator 4.11.3
# Authors: Gaurav, Devashish Yadav Y13 Hall-2

from PyQt4 import QtCore, QtGui
import MySQLdb
import datetime
from openpyxl import Workbook
logs = open('juicecenter.log','a')
data = []
wb = Workbook()
global lastamount
try:
	db = MySQLdb.connect("localhost","root","tiger","juicecenter")
	cursor = db.cursor()
except MySQLdb.Error as e:
	logs.write(time.asctime( time.localtime(time.time()) ) + " : "+str(e)+"\n")


try:
    _fromUtf8 = QtCore.QString.fromUtf8
except AttributeError:
    def _fromUtf8(s):
        return s

try:
    _encoding = QtGui.QApplication.UnicodeUTF8
    def _translate(context, text, disambig):
        return QtGui.QApplication.translate(context, text, disambig, _encoding)
except AttributeError:
    def _translate(context, text, disambig):
        return QtGui.QApplication.translate(context, text, disambig)

#########################################

class Ui_Admin_corner(object):
    def setupUi(self, Admin_corner):
        Admin_corner.setObjectName(_fromUtf8("Admin_corner"))
        Admin_corner.resize(800, 600)
        self.centralwidget = QtGui.QWidget(Admin_corner)
        self.centralwidget.setObjectName(_fromUtf8("centralwidget"))
        self.label = QtGui.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(160, 110, 41, 21))
        self.label.setObjectName(_fromUtf8("label"))
        self.label_2 = QtGui.QLabel(self.centralwidget)
        self.label_2.setGeometry(QtCore.QRect(380, 110, 67, 17))
        self.label_2.setObjectName(_fromUtf8("label_2"))
        self.pushButton = QtGui.QPushButton(self.centralwidget)
        self.pushButton.setGeometry(QtCore.QRect(250, 240, 121, 27))
        self.pushButton.setObjectName(_fromUtf8("pushButton"))
        self.pushButton_2 = QtGui.QPushButton(self.centralwidget)
        self.pushButton_2.setGeometry(QtCore.QRect(550, 20, 99, 27))
        self.pushButton_2.setObjectName(_fromUtf8("pushButton_2"))
        self.label_3 = QtGui.QLabel(self.centralwidget)
        self.label_3.setGeometry(QtCore.QRect(50, 310, 700, 31))
        self.label_3.setObjectName(_fromUtf8("label_3"))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.label_3.setStyleSheet(_fromUtf8("color:red;"))
        self.label_3.setFont(font)
        self.lineEdit = QtGui.QLineEdit(self.centralwidget)
        self.lineEdit.setGeometry(QtCore.QRect(130, 170, 113, 27))
        self.lineEdit.setObjectName(_fromUtf8("lineEdit"))
        self.lineEdit_2 = QtGui.QLineEdit(self.centralwidget)
        self.lineEdit_2.setGeometry(QtCore.QRect(360, 170, 113, 27))
        self.lineEdit_2.setText(_fromUtf8(""))
        self.lineEdit_2.setObjectName(_fromUtf8("lineEdit_2"))
        self.label_4 = QtGui.QLabel(self.centralwidget)
        self.label_4.setGeometry(QtCore.QRect(140, 140, 91, 20))
        self.label_4.setObjectName(_fromUtf8("label_4"))
        self.label_5 = QtGui.QLabel(self.centralwidget)
        self.label_5.setGeometry(QtCore.QRect(370, 140, 91, 20))
        self.label_5.setObjectName(_fromUtf8("label_5"))
        self.label_6 = QtGui.QLabel(self.centralwidget)
        self.label_6.setGeometry(QtCore.QRect(250, 40, 131, 31))
        self.label_6.setObjectName(_fromUtf8("label_6"))
        Admin_corner.setCentralWidget(self.centralwidget)
        self.menubar = QtGui.QMenuBar(Admin_corner)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 800, 25))
        self.menubar.setObjectName(_fromUtf8("menubar"))
        Admin_corner.setMenuBar(self.menubar)
        self.statusbar = QtGui.QStatusBar(Admin_corner)
        self.statusbar.setObjectName(_fromUtf8("statusbar"))
        Admin_corner.setStatusBar(self.statusbar)

        self.retranslateUi(Admin_corner)
        QtCore.QObject.connect(self.pushButton, QtCore.SIGNAL(_fromUtf8("clicked()")), self.trigger)
        QtCore.QObject.connect(self.pushButton_2, QtCore.SIGNAL(_fromUtf8("clicked()")), self.prev)
        QtCore.QMetaObject.connectSlotsByName(Admin_corner)

    def trigger(self):
        # code to generate excel sheet
	global wb
        start = str(self.lineEdit.text())
        end   = str(self.lineEdit_2.text())
        try:
        	if datetime.datetime.strptime(start, '%d-%m-%Y') <= datetime.datetime.strptime(end, '%d-%m-%Y'):
			try:
				cursor.execute("SELECT * FROM students")
				all_list = cursor.fetchall()
				ws = wb.active
				ws.title = "Hall2 "+start+" to "+end
				ws['A1'] = "S_No"
				ws['B1'] = "rollno" 
				ws['C1'] = "Name"
				ws['D1'] = "icecream"
				ws['E1'] = "juice"
				ws['F1'] = "total"
				for i in range(len(all_list)):
					sql = "SELECT * FROM s"+str(all_list[i][2])+" where date_on>='"+str(datetime.datetime.strptime(start, "%d-%m-%Y").strftime("%Y-%m-%d"))+"' and date_on<='"+str(datetime.datetime.strptime(end, "%d-%m-%Y").strftime("%Y-%m-%d"))+" 23:59:59';"
					cursor.execute(sql)
					temp = cursor.fetchall()
					self.label_3.setText(str(((i*1.0)/len(all_list)))*100+"\% Created")
					if(len(temp)>0):
						total_ice = sum(zip(*temp)[2])  
						total_juice = sum(zip(*temp)[3])  
						total_dues = sum(zip(*temp)[4])
						ws['A'+ str(i+2)] = all_list[i][0]
						ws['B'+ str(i+2)] = all_list[i][1] 
						ws['C'+ str(i+2)] = all_list[i][2]
						ws['D'+ str(i+2)] = total_juice
						ws['E'+ str(i+2)] = total_ice
						ws['F'+ str(i+2)] = total_dues
					else:
						ws['A'+ str(i+2)] = all_list[i][0]
						ws['B'+ str(i+2)] = all_list[i][1] 
						ws['C'+ str(i+2)] = all_list[i][2]
						ws['D'+ str(i+2)] = 0
						ws['E'+ str(i+2)] = 0
						ws['F'+ str(i+2)] = 0
				wb.save(start+"-"+end+"_Juice_dues.xlsx")
				self.label_3.setText("Excel successfully created.")

			except MySQLdb.Error as e:
				self.label_3.setText("Error in database. Make sure MySQL server is running")
				print "sql error: "+str(e)
			except Exception as e:
				print e
        	else:
        		self.label_3.setText("Start Date should be less than or equal to end date")

        except Exception as e:
        	print "Enter in the given format only." + str(e)
    		self.label_3.setText("Enter in the given format only.")



    def prev(self):
        ui4.setupUi(MainWindow)    
        

    def retranslateUi(self, Admin_corner):
    	MainWindow.setWindowTitle(_translate("MainWindow", "JuiceCentre 1.0", None))
        self.label.setText(_translate("Admin_corner", "From", None))
        self.label_2.setText(_translate("Admin_corner", "To", None))
        self.pushButton.setText(_translate("Admin_corner", " Generate Excel", None))
        self.pushButton_2.setText(_translate("Admin_corner", "Back", None))
        self.label_3.setText(_translate("Admin_corner", "", None))
        self.label_4.setText(_translate("Admin_corner", "dd-mm-yyyy", None))
        self.label_5.setText(_translate("Admin_corner", "dd-mm-yyyy", None))
        self.label_6.setText(_translate("Admin_corner", "Create Excel", None))


###################################

class Ui_Chooser_window(object):
    def setupUi(self, Chooser_window):
        Chooser_window.setObjectName(_fromUtf8("Chooser_window"))
        Chooser_window.resize(800, 600)
        self.centralwidget = QtGui.QWidget(Chooser_window)
        self.centralwidget.setObjectName(_fromUtf8("centralwidget"))
        self.pushButton = QtGui.QPushButton(self.centralwidget)
        self.pushButton.setGeometry(QtCore.QRect(238, 130, 131, 31))
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(False)
        font.setWeight(50)
        self.pushButton.setFont(font)
        self.pushButton.setObjectName(_fromUtf8("pushButton"))
        self.pushButton_2 = QtGui.QPushButton(self.centralwidget)
        self.pushButton_2.setGeometry(QtCore.QRect(238, 200, 131, 31))
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(False)
        font.setWeight(50)
        self.pushButton_2.setFont(font)
        self.pushButton_2.setObjectName(_fromUtf8("pushButton_2"))
        Chooser_window.setCentralWidget(self.centralwidget)
        self.menubar = QtGui.QMenuBar(Chooser_window)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 800, 25))
        self.menubar.setObjectName(_fromUtf8("menubar"))
        Chooser_window.setMenuBar(self.menubar)
        self.statusbar = QtGui.QStatusBar(Chooser_window)
        self.statusbar.setObjectName(_fromUtf8("statusbar"))
        Chooser_window.setStatusBar(self.statusbar)

        self.retranslateUi(Chooser_window)
        QtCore.QObject.connect(self.pushButton, QtCore.SIGNAL(_fromUtf8("clicked()")), self.juice)
        QtCore.QObject.connect(self.pushButton_2, QtCore.SIGNAL(_fromUtf8("clicked()")), self.admin)
        QtCore.QMetaObject.connectSlotsByName(Chooser_window)

    def juice(self):
        ui1.setupUi(MainWindow)

    def admin(self):
        ui5.setupUi(MainWindow)    
        
    def retranslateUi(self, Chooser_window):
    	MainWindow.setWindowTitle(_translate("MainWindow", "JuiceCentre 1.0", None))
        self.pushButton.setText(_translate("Chooser_window", "Juice Entry", None))
        self.pushButton_2.setText(_translate("Chooser_window", "Admin Corner", None))

####################################

class Ui_login_page(object):
    def setupUi(self, login_page):
        login_page.setObjectName(_fromUtf8("login_page"))
        login_page.resize(800, 600)
        self.centralwidget = QtGui.QWidget(login_page)
        self.centralwidget.setObjectName(_fromUtf8("centralwidget"))
        self.pushButton_2 = QtGui.QPushButton(self.centralwidget)
        self.pushButton_2.setGeometry(QtCore.QRect(370, 220, 99, 27))
        self.pushButton_2.setObjectName(_fromUtf8("pushButton_2"))
        self.label = QtGui.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(120, 120, 71, 31))
        self.label.setObjectName(_fromUtf8("label"))
        self.label_2 = QtGui.QLabel(self.centralwidget)
        self.label_2.setGeometry(QtCore.QRect(120, 160, 81, 31))
        self.label_2.setObjectName(_fromUtf8("label_2"))
        self.label_3 = QtGui.QLabel(self.centralwidget)
        self.label_3.setGeometry(QtCore.QRect(50, 290, 700, 31))
        self.label_3.setObjectName(_fromUtf8("label_3"))
	font = QtGui.QFont()
        font.setPointSize(14)
        self.label_3.setStyleSheet(_fromUtf8("color:red;"))
        self.label_3.setFont(font)
        self.lineEdit = QtGui.QLineEdit(self.centralwidget)
        self.lineEdit.setGeometry(QtCore.QRect(270, 120, 201, 27))
        self.lineEdit.setObjectName(_fromUtf8("lineEdit"))
        self.lineEdit_2 = QtGui.QLineEdit(self.centralwidget)
        self.lineEdit_2.setGeometry(QtCore.QRect(270, 160, 201, 27))
        self.lineEdit_2.setObjectName(_fromUtf8("lineEdit_2"))
        self.lineEdit_2.setEchoMode(QtGui.QLineEdit.Password)
        login_page.setCentralWidget(self.centralwidget)
        self.menubar = QtGui.QMenuBar(login_page)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 800, 25))
        self.menubar.setObjectName(_fromUtf8("menubar"))
        login_page.setMenuBar(self.menubar)
        self.statusbar = QtGui.QStatusBar(login_page)
        self.statusbar.setObjectName(_fromUtf8("statusbar"))
        login_page.setStatusBar(self.statusbar)

        self.retranslateUi(login_page)
        QtCore.QObject.connect(self.pushButton_2, QtCore.SIGNAL(_fromUtf8("clicked()")), self.login_action)
	QtCore.QObject.connect(self.lineEdit, QtCore.SIGNAL(_fromUtf8("returnPressed()")), self.pushButton_2.click)
	QtCore.QObject.connect(self.lineEdit_2, QtCore.SIGNAL(_fromUtf8("returnPressed()")), self.pushButton_2.click)
        QtCore.QMetaObject.connectSlotsByName(login_page)
        login_page.setTabOrder(self.lineEdit, self.lineEdit_2)
        login_page.setTabOrder(self.lineEdit_2, self.pushButton_2)

    def login_action(self):
        login = str(self.lineEdit.text())
        pswrd = str(self.lineEdit_2.text())
	try:
		cursor.execute("SELECT * FROM login")
		user_data = cursor.fetchone()
		if login == str(user_data[0]) and pswrd == str(user_data[1]): 
		    ui4.setupUi(MainWindow)
		else :
		    self.label_3.setText(_translate("login_page", "Wrong login credintials", None))
	except MySQLdb.Error as e:
		self.label_3.setText(_translate("login_page", "Database Error", None))
	except Exception,e:
		self.label_3.setText(_translate("login_page", "Exception:"+str(e), None))

    def retranslateUi(self, login_page):
        MainWindow.setWindowTitle(_translate("MainWindow", "JuiceCentre 1.0", None))
        self.pushButton_2.setText(_translate("login_page", "Go", None))
        self.label.setText(_translate("login_page", "Login:", None))
        self.label_2.setText(_translate("login_page", "Password:", None))
        self.label_3.setText(_translate("login_page", "", None))        

########################################
class Ui_MainWindow(object):

    def setupUi(self, MainWindow):
        MainWindow.setObjectName(_fromUtf8("MainWindow"))
        MainWindow.setEnabled(True)
        MainWindow.resize(800, 600)
        sizePolicy = QtGui.QSizePolicy(QtGui.QSizePolicy.Fixed, QtGui.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(MainWindow.sizePolicy().hasHeightForWidth())
        MainWindow.setSizePolicy(sizePolicy)
        MainWindow.setMinimumSize(QtCore.QSize(800, 600))
        self.centralWidget = QtGui.QWidget(MainWindow)
        self.centralWidget.setObjectName(_fromUtf8("centralWidget"))
        self.lineEdit = QtGui.QLineEdit(self.centralWidget)
        self.lineEdit.setGeometry(QtCore.QRect(112, 159, 316, 50))
        self.lineEdit.setMinimumSize(QtCore.QSize(231, 35))
        self.lineEdit.setStyleSheet(_fromUtf8("font: 26pt \"Ubuntu\";\n"
"text-align: center;"))
        self.lineEdit.setText(_fromUtf8(""))
        self.lineEdit.setObjectName(_fromUtf8("lineEdit"))
        self.pushButton = QtGui.QPushButton(self.centralWidget)
        self.pushButton.setGeometry(QtCore.QRect(434, 157, 85, 51))
        self.pushButton.setStyleSheet(_fromUtf8("font: 63 16pt \"Ubuntu\";"))
        self.pushButton.setObjectName(_fromUtf8("pushButton"))
        self.label = QtGui.QLabel(self.centralWidget)
        self.label.setGeometry(QtCore.QRect(111, 111, 166, 32))
        self.label.setStyleSheet(_fromUtf8("font: 75 20pt \"Ubuntu\";\n"
"color: rgb(0, 85, 255)"))
        self.label.setObjectName(_fromUtf8("label"))
        self.label_2 = QtGui.QLabel(self.centralWidget)
        self.label_2.setGeometry(QtCore.QRect(110, 260, 311, 31))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.label_2.setFont(font)
        self.label_2.setText(_fromUtf8(""))
        self.label_2.setStyleSheet(_fromUtf8("color:red;"))
        self.label_2.setObjectName(_fromUtf8("label_2"))
	self.label_last = QtGui.QLabel(self.centralWidget)
        self.label_last.setGeometry(QtCore.QRect(30, 330, 700, 31))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_last.setFont(font)
        self.label_last.setText(_fromUtf8(""))
        self.label_last.setStyleSheet(_fromUtf8("color:black;"))
        self.label_last.setObjectName(_fromUtf8("label_2"))
        MainWindow.setCentralWidget(self.centralWidget)
        self.menuBar = QtGui.QMenuBar(MainWindow)
        self.menuBar.setGeometry(QtCore.QRect(0, 0, 800, 25))
        self.menuBar.setObjectName(_fromUtf8("menuBar"))
        self.menuJuiceCentre = QtGui.QMenu(self.menuBar)
        self.menuJuiceCentre.setObjectName(_fromUtf8("menuJuiceCentre"))
        MainWindow.setMenuBar(self.menuBar)
        self.mainToolBar = QtGui.QToolBar(MainWindow)
        self.mainToolBar.setObjectName(_fromUtf8("mainToolBar"))
        MainWindow.addToolBar(QtCore.Qt.TopToolBarArea, self.mainToolBar)
        self.statusBar = QtGui.QStatusBar(MainWindow)
        self.statusBar.setObjectName(_fromUtf8("statusBar"))
        MainWindow.setStatusBar(self.statusBar)
        self.menuBar.addAction(self.menuJuiceCentre.menuAction())

        self.retranslateUi(MainWindow)
        QtCore.QObject.connect(self.pushButton, QtCore.SIGNAL(_fromUtf8("clicked()")), self.rollnumber)
        QtCore.QObject.connect(self.lineEdit, QtCore.SIGNAL(_fromUtf8("returnPressed()")), self.pushButton.click)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
	global lastamount
        MainWindow.setWindowTitle(_translate("MainWindow", "JuiceCentre 1.0", None))
        self.pushButton.setText(_translate("MainWindow", "ENTER", None))
        self.label.setText(_translate("MainWindow", "Enter Roll No:", None))
        self.menuJuiceCentre.setTitle(_translate("MainWindow", "JuiceCentre", None))
	if len(data)!=0:
		self.label_last.setText(_translate("MainWindow", "Last Transaction:      "+str(data[0][2])+" "+str(data[0][1])+"             Amount: Rs "+ str(lastamount), None))
	
    def rollnumber(self):
		rollno = str(self.lineEdit.text())
		global data
		#print rollno
		if rollno =="":
			rollno="0"			
		sqlq = "SELECT * from students where rollno="+str(rollno)
		cursor.execute(sqlq)
		data = cursor.fetchall()
		if len(data)==0:
			print "Invalid Roll No."
			self.label_2.setText("Invalid Roll No.")
		else:
			ui2.setupUi(MainWindow)

##########################################

class Ui_StudentPage(object):
    def setupUi(self, StudentPage):
        StudentPage.setObjectName(_fromUtf8("StudentPage"))
        StudentPage.resize(800, 600)
        sizePolicy = QtGui.QSizePolicy(QtGui.QSizePolicy.Fixed, QtGui.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(50)
        sizePolicy.setVerticalStretch(50)
        sizePolicy.setHeightForWidth(StudentPage.sizePolicy().hasHeightForWidth())
        StudentPage.setSizePolicy(sizePolicy)
        StudentPage.setMinimumSize(QtCore.QSize(800, 600))
        self.centralwidget = QtGui.QWidget(StudentPage)
        self.centralwidget.setObjectName(_fromUtf8("centralwidget"))
        self.lineEdit = QtGui.QLineEdit(self.centralwidget)
        self.lineEdit.setGeometry(QtCore.QRect(171, 221, 316, 50))
        self.lineEdit.setStyleSheet(_fromUtf8("font: 26pt \"Ubuntu\";"))
        self.lineEdit.setText(_fromUtf8("0"))
        self.lineEdit.setObjectName(_fromUtf8("lineEdit"))
        self.pushButton = QtGui.QPushButton(self.centralwidget)
        self.pushButton.setGeometry(QtCore.QRect(510, 280, 101, 51))
        self.pushButton.setObjectName(_fromUtf8("pushButton"))
        self.label = QtGui.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(170, 170, 181, 41))
        self.label.setStyleSheet(_fromUtf8("font: 75 20pt \"Ubuntu\";\n"
"color : rgb(0, 85, 255);"))
        self.label.setObjectName(_fromUtf8("label"))
        self.rollno = QtGui.QLabel(self.centralwidget)
        self.rollno.setGeometry(QtCore.QRect(171, 59, 379, 41))
        self.rollno.setText(_fromUtf8(""))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.rollno.setFont(font)
        self.rollno.setObjectName(_fromUtf8("rollno"))
        self.image = QtGui.QLabel(self.centralwidget)
        self.image.setObjectName(_fromUtf8("image"))
        self.image.setGeometry(QtCore.QRect(10, 10, 141, 150))
        self.image.setPixmap(QtGui.QPixmap("../images/"+str(data[0][2])+"_0.jpg"))
        self.image.setScaledContents(True);
        self.name = QtGui.QLabel(self.centralwidget)
        self.name.setGeometry(QtCore.QRect(171, 11, 379, 42))
        self.name.setText(_fromUtf8(""))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.name.setFont(font)
        self.name.setObjectName(_fromUtf8("name"))
        self.pushButton_2 = QtGui.QPushButton(self.centralwidget)
        self.pushButton_2.setGeometry(QtCore.QRect(558, 10, 111, 41))
        self.pushButton_2.setObjectName(_fromUtf8("pushButton_2"))
        self.label_2 = QtGui.QLabel(self.centralwidget)
        self.label_2.setGeometry(QtCore.QRect(30, 240, 121, 21))
        font = QtGui.QFont()
        font.setPointSize(20)
        self.label_2.setFont(font)
        self.label_2.setObjectName(_fromUtf8("label_2"))
        self.label_3 = QtGui.QLabel(self.centralwidget)
        self.label_3.setGeometry(QtCore.QRect(50, 290, 81, 31))
        font = QtGui.QFont()
        font.setPointSize(20)
        self.label_3.setFont(font)
        self.label_3.setObjectName(_fromUtf8("label_3"))
        self.label_4 = QtGui.QLabel(self.centralwidget)
        self.label_4.setGeometry(QtCore.QRect(30, 340, 700, 31))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.label_4.setStyleSheet(_fromUtf8("color:red;"))
        self.label_4.setFont(font)
        self.label_4.setObjectName(_fromUtf8("label_4"))
        self.lineEdit_2 = QtGui.QLineEdit(self.centralwidget)
        self.lineEdit_2.setGeometry(QtCore.QRect(170, 280, 316, 50))
        self.lineEdit_2.setStyleSheet(_fromUtf8("font: 26pt \"Ubuntu\";"))
        self.lineEdit_2.setText(_fromUtf8("0"))
        self.lineEdit_2.setObjectName(_fromUtf8("lineEdit_2"))
        StudentPage.setCentralWidget(self.centralwidget)
        self.menubar = QtGui.QMenuBar(StudentPage)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 800, 25))
        self.menubar.setObjectName(_fromUtf8("menubar"))
        StudentPage.setMenuBar(self.menubar)
        self.statusbar = QtGui.QStatusBar(StudentPage)
        self.statusbar.setObjectName(_fromUtf8("statusbar"))
        StudentPage.setStatusBar(self.statusbar)
        self.toolBar = QtGui.QToolBar(StudentPage)
        self.toolBar.setObjectName(_fromUtf8("toolBar"))
        StudentPage.addToolBar(QtCore.Qt.TopToolBarArea, self.toolBar)
        self.toolBar_2 = QtGui.QToolBar(StudentPage)
        self.toolBar_2.setObjectName(_fromUtf8("toolBar_2"))
        StudentPage.addToolBar(QtCore.Qt.TopToolBarArea, self.toolBar_2)

        self.retranslateUi(StudentPage)
        QtCore.QObject.connect(self.pushButton, QtCore.SIGNAL(_fromUtf8("clicked()")), self.store)
        QtCore.QObject.connect(self.pushButton_2, QtCore.SIGNAL(_fromUtf8("clicked()")), self.cancel)
        QtCore.QObject.connect(self.lineEdit, QtCore.SIGNAL(_fromUtf8("returnPressed()")), self.pushButton.click)
        QtCore.QObject.connect(self.lineEdit_2, QtCore.SIGNAL(_fromUtf8("returnPressed()")), self.pushButton.click)
        QtCore.QMetaObject.connectSlotsByName(StudentPage)
        StudentPage.setTabOrder(self.lineEdit, self.lineEdit_2)
        
    def retranslateUi(self, StudentPage):
    	global data
        StudentPage.setWindowTitle(_translate("StudentPage", "JuiceCentre 1.0", None))
        self.pushButton.setText(_translate("StudentPage", "ENTER", None))
        self.label.setText(_translate("StudentPage", "Enter Amount:", None))
        self.pushButton_2.setText(_translate("StudentPage", "CANCEL", None))
        self.label_2.setText(_translate("StudentPage", "Ice-Cream", None))
        self.label_3.setText(_translate("StudentPage", "Juice", None))
        self.toolBar.setWindowTitle(_translate("StudentPage", "toolBar", None))
        self.toolBar_2.setWindowTitle(_translate("StudentPage", "toolBar_2", None))
    	self.name.setText(_translate("StudentPage", str(data[0][1]), None))
    	self.rollno.setText(_translate("StudentPage", str(data[0][2]), None))

    def cancel(self):
    	ui1.setupUi(MainWindow)

    def store(self):
        icecream = str(self.lineEdit.text())
        juice    = str(self.lineEdit_2.text())
        if len(juice)==0:
        	juice = 0
        if len(icecream)==0:
        	icecream = 0
    	try:
    		icecream = int(icecream)
    		juice = int(juice)
    		#print "done icecream= %s juice=%s" %(icecream,juice)
    		if(icecream>0 or juice>0):
	    		try:
	    			cursor.execute ("""
	   							INSERT into s"""+str(data[0][2])+"""
								(icecream,juice,amount) values ("%s","%s","%s")
								""", (icecream,juice,juice+icecream))
	    			db.commit()
				global lastamount
				lastamount = icecream+juice
	    			ui1.setupUi(MainWindow)
	    		except MySQLdb.Error as e:
	    			db.rollback()
	    			self.label_4.setText("Database Error. Press cancel and try again.")
	    			print "Database Error. Press cancel and try again. \n error: %s"%e
	    	else:
	    		self.label_4.setText("Please Valid (greater than 0) input and hit ENTER or press CANCEL")
	    		print "Please Valid (greater than 0) input and hit ENTER or press CANCEL"
    	except Exception, e:
    		self.label_4.setText("Enter a vaild number")
    		print "Please enter valid number. %s"%e

if __name__ == "__main__":
    import sys
    app = QtGui.QApplication(sys.argv)
    MainWindow = QtGui.QMainWindow()
    palette = QtGui.QPalette()
    palette.setColor(QtGui.QPalette.Background,QtCore.Qt.white)
    MainWindow.setPalette(palette)
    ui1 = Ui_MainWindow()
    ui2 = Ui_StudentPage()
    ui3 = Ui_login_page()
    ui4 = Ui_Chooser_window()
    ui5 = Ui_Admin_corner()

    ui3.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())

